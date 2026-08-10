"""The Dashboard's Excel export.

A real workbook, not a CSV renamed — so the tests open what the endpoint
produced and read the cells back, rather than trusting the bytes. The two things
worth guarding are that dates land as dates (the reason for xlsx at all) and
that the export cannot become the one door that ignores project access.
"""
from __future__ import annotations

import datetime as dt
from io import BytesIO

import pytest
from openpyxl import load_workbook
from rest_framework.test import APIClient

from apps.accounts.models import Role, User
from apps.audit.models import AuditAction, AuditLog
from apps.workspaces.models import WorkspaceMember, WorkspaceProject, WorkspaceProjectMember

EXPORT = "/api/workspace-projects/export.xlsx"
WORKSPACE = "amazon-usa"
HEADERS = ["Project name", "Assigned to", "Assigned by", "Due date", "Status", "Priority"]


def _sheet(response):
    return load_workbook(BytesIO(response.content)).active


def _rows(response) -> list[dict]:
    sheet = _sheet(response)
    values = list(sheet.values)
    return [dict(zip(values[0], row)) for row in values[1:]]


@pytest.fixture
def executive_role(db) -> Role:
    return Role.objects.create(name="Executive")


def _exec(username: str, role: Role, full: str) -> User:
    first, last = full.split(" ", 1)
    user = User.objects.create_user(username=username, email=f"{username}@kos.test",
                                    password="pw-exec-1234", first_name=first, last_name=last)
    user.roles.add(role)
    WorkspaceMember.objects.create(user=user, workspace=WORKSPACE, access=WorkspaceMember.EDIT)
    return user


def _client(user: User) -> APIClient:
    client = APIClient()
    client.force_authenticate(user)
    return client


# --------------------------------------------------------------------------- #
# Shape and content
# --------------------------------------------------------------------------- #
@pytest.mark.django_db
def test_the_sheet_has_the_six_requested_columns(auth_client):
    WorkspaceProject.objects.create(workspace=WORKSPACE, name="Neem Oil 2026")
    r = auth_client.get(EXPORT)
    assert r.status_code == 200, r.content[:200]
    assert r["Content-Type"].endswith("spreadsheetml.sheet")
    assert "attachment; filename=" in r["Content-Disposition"]
    assert list(_sheet(r).values)[0] == tuple(HEADERS)


@pytest.mark.django_db
def test_the_due_date_is_a_real_date_not_text(auth_client):
    """The whole reason for xlsx over CSV: a date that sorts as a date."""
    WorkspaceProject.objects.create(
        workspace=WORKSPACE, name="Label revision",
        start_at=dt.datetime(2026, 8, 1, 9, tzinfo=dt.timezone.utc),
        end_at=dt.datetime(2026, 8, 20, 17, tzinfo=dt.timezone.utc),
    )
    due = _rows(auth_client.get(EXPORT))[0]["Due date"]
    assert isinstance(due, (dt.datetime, dt.date)), f"got {type(due)}"
    assert (due.date() if isinstance(due, dt.datetime) else due) == dt.date(2026, 8, 20)


@pytest.mark.django_db
def test_a_project_with_no_end_leaves_the_cell_empty(auth_client):
    WorkspaceProject.objects.create(workspace=WORKSPACE, name="Open ended")
    assert _rows(auth_client.get(EXPORT))[0]["Due date"] is None


@pytest.mark.django_db
def test_assigned_to_is_the_roster_and_assigned_by_is_who_put_them_there(
        auth_client, admin_user, executive_role):
    alice = _exec("alice", executive_role, "Alice Rao")
    bob = _exec("bob", executive_role, "Bob Menon")
    project = WorkspaceProject.objects.create(workspace=WORKSPACE, name="Neem Oil 2026", created_by=alice)
    WorkspaceProjectMember.objects.create(project=project, user=bob, added_by=alice)

    row = _rows(auth_client.get(EXPORT))[0]
    assert row["Assigned to"] == "Bob Menon"
    assert row["Assigned by"] == "Alice Rao"


@pytest.mark.django_db
def test_several_members_are_listed_on_the_one_row(auth_client, executive_role):
    alice = _exec("alice", executive_role, "Alice Rao")
    bob = _exec("bob", executive_role, "Bob Menon")
    project = WorkspaceProject.objects.create(workspace=WORKSPACE, name="Neem Oil 2026", created_by=alice)
    for user in (alice, bob):
        WorkspaceProjectMember.objects.create(project=project, user=user, added_by=alice)

    assert _rows(auth_client.get(EXPORT))[0]["Assigned to"] == "Alice Rao; Bob Menon"


@pytest.mark.django_db
def test_with_no_roster_assigned_by_falls_back_to_the_creator(auth_client, executive_role):
    alice = _exec("alice", executive_role, "Alice Rao")
    WorkspaceProject.objects.create(workspace=WORKSPACE, name="Unassigned", created_by=alice)
    row = _rows(auth_client.get(EXPORT))[0]
    # A genuinely empty cell rather than the text "" — it filters as blank,
    # which is what someone sorting for unassigned work will reach for.
    assert row["Assigned to"] is None
    assert row["Assigned by"] == "Alice Rao"


@pytest.mark.django_db
@pytest.mark.parametrize("days,status,priority", [
    (-3, "Overdue", "High"),
    (0.5, "Ending soon", "Medium"),
    (30, "In progress", "Low"),
])
def test_status_and_the_derived_priority(auth_client, days, status, priority):
    now = dt.datetime.now(dt.timezone.utc)
    WorkspaceProject.objects.create(
        workspace=WORKSPACE, name="Timed", start_at=now - dt.timedelta(days=10),
        end_at=now + dt.timedelta(days=days),
    )
    row = _rows(auth_client.get(EXPORT))[0]
    assert row["Status"] == status
    # Priority is derived from the schedule — it has no field of its own.
    assert row["Priority"] == priority


@pytest.mark.django_db
def test_the_priority_header_says_it_is_derived(auth_client):
    """A spreadsheet outlives the conversation that produced it, so the caveat
    has to travel inside the file."""
    WorkspaceProject.objects.create(workspace=WORKSPACE, name="Neem Oil 2026")
    sheet = _sheet(auth_client.get(EXPORT))
    comment = sheet.cell(row=1, column=HEADERS.index("Priority") + 1).comment
    assert comment is not None and "Derived from the schedule" in comment.text


# --------------------------------------------------------------------------- #
# Access
# --------------------------------------------------------------------------- #
@pytest.mark.django_db
def test_a_non_admin_cannot_export(executive_role):
    alice = _exec("alice", executive_role, "Alice Rao")
    WorkspaceProject.objects.create(workspace=WORKSPACE, name="Neem Oil 2026")
    assert _client(alice).get(EXPORT).status_code == 403


@pytest.mark.django_db
def test_anonymous_callers_are_turned_away(api_client):
    assert api_client.get(EXPORT).status_code == 401


@pytest.mark.django_db
def test_archived_workspaces_stay_out_of_the_export(auth_client):
    from apps.workspaces.models import Workspace

    WorkspaceProject.objects.create(workspace=WORKSPACE, name="Live one")
    Workspace.objects.create(key="gone", label="Gone", archived_at=dt.datetime.now(dt.timezone.utc))
    WorkspaceProject.objects.create(workspace="gone", name="Archived one")

    names = [r["Project name"] for r in _rows(auth_client.get(EXPORT))]
    assert names == ["Live one"]


@pytest.mark.django_db
def test_deleted_projects_stay_out_of_the_export(auth_client):
    WorkspaceProject.objects.create(workspace=WORKSPACE, name="Live one")
    WorkspaceProject.objects.create(
        workspace=WORKSPACE, name="Binned", deleted_at=dt.datetime.now(dt.timezone.utc))
    names = [r["Project name"] for r in _rows(auth_client.get(EXPORT))]
    assert names == ["Live one"]


@pytest.mark.django_db
def test_the_export_is_audited(auth_client):
    WorkspaceProject.objects.create(workspace=WORKSPACE, name="Neem Oil 2026")
    auth_client.get(EXPORT)
    entry = AuditLog.objects.filter(action=AuditAction.EXPORT).latest("id")
    assert entry.new_value["kind"] == "dashboard_xlsx" and entry.new_value["rows"] == 1
