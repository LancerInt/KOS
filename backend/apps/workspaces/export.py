"""Excel export of the Dashboard's project list.

An administrators-only ``.xlsx`` of every project the dashboard can show, for
the reporting that happens outside KOS — a board pack, a review meeting, a
column someone wants to sort three different ways.

A real workbook rather than a CSV renamed: dates are written as dates, so they
sort and filter as dates instead of as the text "13/08/2026". The Reports tab's
CSV export is unchanged and stays where it is.
"""
from __future__ import annotations

from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# The dashboard's own wording, so the spreadsheet and the screen agree.
STATUS_LABELS = {
    "none": "No duration",
    "active": "In progress",
    "ending_soon": "Ending soon",
    "due": "Overdue",
    "completed": "Completed",
}

#: Priority is **derived from the schedule**, not chosen by anyone: a project
#: carries no priority field. Overdue is the only thing that reads as urgent
#: without someone having said so, and "ending soon" is the only warning the
#: data supports. Everything else is Low by default rather than by judgement —
#: see the caveat in the sheet's own header comment.
PRIORITY_BY_STATUS = {"due": "High", "ending_soon": "Medium"}
DEFAULT_PRIORITY = "Low"

COLUMNS = [
    ("Project name", 38),
    ("Assigned to", 30),
    ("Assigned by", 22),
    ("Due date", 13),
    ("Status", 15),
    ("Priority", 10),
]

HEADER_FILL = PatternFill("solid", fgColor="0F7A8B")   # the brand teal
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)


def _person(user) -> str:
    if user is None:
        return ""
    return user.get_full_name() or user.username


def project_rows(projects) -> list[dict]:
    """One row per project, in the order the queryset supplies.

    ``Assigned to`` is the project's member roster — the people it was actually
    given to. ``Assigned by`` is whoever put them there; with no roster it falls
    back to whoever created the project, which is the closest true answer to
    "who assigned this" when nobody has been assigned yet.
    """
    rows = []
    for project in projects:
        members = list(project.members.all())
        assigned_to = "; ".join(sorted(_person(m.user) for m in members))
        assigners = sorted({_person(m.added_by) for m in members if m.added_by_id})
        assigned_by = "; ".join(assigners) if assigners else _person(project.created_by)
        status = (project.duration_state() or {}).get("status", "none")

        rows.append({
            "Project name": project.name,
            "Assigned to": assigned_to,
            "Assigned by": assigned_by,
            # A date object, not a string — this is the whole point of xlsx.
            "Due date": timezone.localtime(project.end_at).date() if project.end_at else None,
            # The schedule state, which is what Status has always meant on the
            # dashboard and what Priority below is derived from. Blocked /
            # Needs decision are a separate axis and are deliberately not folded
            # in here — one column cannot carry two states without losing one.
            "Status": STATUS_LABELS.get(status, status),
            "Priority": PRIORITY_BY_STATUS.get(status, DEFAULT_PRIORITY),
        })
    return rows


def build_workbook(rows: list[dict]) -> bytes:
    """The rows as a formatted .xlsx, returned as bytes."""
    book = Workbook()
    sheet = book.active
    sheet.title = "Projects"

    headers = [name for name, _ in COLUMNS]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 20

    for row in rows:
        sheet.append([row[name] for name in headers])

    for index, (_, width) in enumerate(COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    due_column = get_column_letter(headers.index("Due date") + 1)
    for cell in sheet[due_column][1:]:            # skip the header
        cell.number_format = "DD/MM/YYYY"

    # Freeze the header and switch on autofilter: the first two things anyone
    # does to a sheet like this by hand.
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(1, len(rows)) + 1}"

    # Priority is derived, and a spreadsheet gets forwarded far from the person
    # who exported it. Say so *in the file*, on the cell that would otherwise be
    # taken as somebody's judgement.
    priority_header = sheet.cell(row=1, column=headers.index("Priority") + 1)
    priority_header.comment = _priority_comment()

    stream = BytesIO()
    book.save(stream)
    return stream.getvalue()


def _priority_comment():
    from openpyxl.comments import Comment

    note = Comment(
        "Derived from the schedule, not set by a person: Overdue = High, "
        "Ending soon = Medium, everything else = Low.",
        "KOS",
    )
    note.width = 320
    note.height = 90
    return note
